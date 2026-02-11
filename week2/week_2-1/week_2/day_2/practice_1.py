# ITP Week 2 Day 2 (In-Class) Practice 1
# 
# You will continue to work on the inventory spreadsheet 
# that you created from yesterday's exercise
# import the appropriate method from the correct module

from openpyxl import load_workbook

# Import the workbook that you created in yesterday's exercise from
wb = load_workbook("week2/week_2-1/week_2/spreadsheets/inventory.xlsx")

# verify what sheet names are available
print("Sheets in workbook:\n" + str(wb.sheetnames))

for sheet in wb:
    print(sheet.title)

# access and store the appropriate worksheet to the variable 'ws'
ws = wb.active

# Print out the cell values for each row
print("\nWorksheet contents:")
for row in list(ws.values):
    print(row)

# Create a new column within that worksheet called order_amount
ws['F1'] = "order_amount"

# save the latest changes
wb.save("week2/week_2-1/week_2/spreadsheets/inventory.xlsx")