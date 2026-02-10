# from openpyxl import Workbook

# wb = Workbook()

# ws = wb.active
# ws.title = "Hello"

# wb.create_sheet("Rugrats")
# wb.create_sheet("First Sheet", 0)

# print(wb.sheetnames)

# ws["A1"] = "Hello"
# print(ws["A1"].value)

# ws.cell(row=2, column=2, value="World")
# print(ws.cell(row=2, column=2).value)  # Reads "World"

# for x in range(1, 101):
#     for y in range(1, 101):
#         ws.cell(row=x, column=y)

# colC = ws["C"]     # entire column C
# row10 = ws[10]     # entire row 10

# for cell in colC:
#     cell.value = "new data"

# for row in ws.iter_rows(min_row=1, max_row=10, min_col=5, max_col=10):
#     for cell in row:
#         print(cell)

# wb.save("Demo.xlsx")
# --------------------------------------------------------------------------------

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "First Sheet"

ws["A1"] = "Hello!"
ws.cell(row=1, column=2, value="There!!")

wb.save("week2/week_2-1/week_2/spreadsheets/Demoday1.xlsx")